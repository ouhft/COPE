#!/usr/bin/python
# coding: utf-8
from datetime import datetime
from openpyxl import Workbook

from django.core.management.base import LabelCommand
from django.utils import timezone

from wp4.samples.utils import WP7Workbook, number_as_str, get_sample_by_barcode
from wp4.samples.models import WP7Record
from wp4.samples.forms import WP7RecordForm


class Command(LabelCommand):
    help = "Load a specified WP7 Export provided spreadsheet of data"

    def handle_label(self, label, **options):
        # Generate some stats for feedback
        start_time = timezone.now()
        end_time = None
        total_rows = 0
        created_count = 0
        update_count = 0

        if label.split(".")[-1] != "xlsx":
            raise Exception("This is not an xlsx file")
        print("load_wp7: Workbook ({0}) is about to load".format(label))

        # Load the xlsx file with the raw data
        workbook = WP7Workbook()
        if not workbook.load_xlsx(label):
            raise Exception("xlsx file failed to load")
        total_rows = workbook.worksheet.max_row
        print("load_wp7: Workbook is now loaded. {0} rows were found".format(total_rows))

        # Iterate through the data, creating or updating records in WP7Record
        for row_index in range(2, total_rows+1):
            row_data = workbook.load_row(row_index)
            # print("DEBUG: load_wp7: Row data for row {0} is loaded: {1}".format(row_index, row_data))

            def cell_value_by_id(column_id=1):
                return workbook.worksheet.cell(row=row_index, column=column_id).value

            def cell_value_by_title(column_name=""):
                return row_data[column_name.lower()]

            # Get or create a record by using the barcode as a key
            barcode = number_as_str(cell_value_by_title("ScannedBarcode"))

            record, created = WP7Record.objects.get_or_create(
                barcode=barcode,
                defaults={
                    'barcode': barcode,
                }
            )
            created_count += 1 if created else 0

            # print("DEBUG: load_wp7: record {0} is created {1} for barcode {2}".format(record, created, barcode))
            # Attempt to match it to an existing sample record
            matched_sample = get_sample_by_barcode(barcode)

            data_form = WP7RecordForm(data={
                'barcode': row_data["scannedbarcode"],
                'box_number': row_data["boxnumber"],
                'position_in_box': row_data["positioninbox"]
            }, instance=record)
            if data_form.is_valid():
                if data_form.has_changed():
                    update_count += 1
                    record = data_form.save()
                record.content_object = matched_sample
                record.save()
                # print("DEBUG: load_wp7: record {0} has content_type {1} for matched_sample {2}".format(
                #     record, record.content_type, matched_sample))
            else:
                print("Form #{0} is INVALID".format(row_index))
                print(data_form.errors)

            # Provide some feedback
            if row_index % 50 == 0:
                print("Row {0} of {1} imported".format(row_index, total_rows))
                print("{0:06.2f}% Complete at {1:%Y-%m-%d %H:%M:%S}".format(
                    (row_index/total_rows)*100,
                    timezone.now()
                ))

        end_time = timezone.now()
        print(
            "Import completed: Started: {0:%Y-%m-%d %H:%M:%S}. ".format(start_time) +
            "Finished: {0:%Y-%m-%d %H:%M:%S}.".format(end_time) +
            "New: {0}. Updated: {1}. Total: {2}".format(created_count, update_count, total_rows)
        )
