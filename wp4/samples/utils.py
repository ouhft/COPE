#!/usr/bin/python
# coding: utf-8
from __future__ import absolute_import, unicode_literals

from django.db import transaction

from openpyxl import Workbook, load_workbook
from io import IOBase

from wp4.locations.models import UNITED_KINGDOM
from .models import Event, BloodSample, UrineSample, PerfusateSample, TissueSample


@transaction.atomic
def create_donor_samples(donor):
    """
    For a given Donor record, create the related events and samples for the protocol
    :param donor: Donor instance
    :return: True, if created, False otherwise
    """

    if donor.retrieval_team.based_at.country != UNITED_KINGDOM:
        person = donor.person
        # No donor samples are taken in the UK
        # ======================================================  BLOOD 1
        blood_event_1 = Event()
        blood_event_1.type = Event.TYPE_BLOOD
        blood_event_1.name = Event.NAME_DONOR_BLOOD1
        blood_event_1.save()

        blood_event_1_sample_1 = BloodSample()
        blood_event_1_sample_1.person = person
        blood_event_1_sample_1.event = blood_event_1
        blood_event_1_sample_1.blood_type = BloodSample.SAMPLE_SST
        blood_event_1_sample_1.save()

        blood_event_1_sample_2 = BloodSample()
        blood_event_1_sample_2.person = person
        blood_event_1_sample_2.event = blood_event_1
        blood_event_1_sample_2.blood_type = BloodSample.SAMPLE_EDSA
        blood_event_1_sample_2.save()

        # ======================================================  URINE 1
        urine_event_1 = Event()
        urine_event_1.type = Event.TYPE_URINE
        urine_event_1.name = Event.NAME_DONOR_URINE1
        urine_event_1.save()

        urine_event_1_sample_1 = UrineSample()
        urine_event_1_sample_1.person = person
        urine_event_1_sample_1.event = urine_event_1
        urine_event_1_sample_1.save()

        # ======================================================  URINE 2
        urine_event_2 = Event()
        urine_event_2.type = Event.TYPE_URINE
        urine_event_2.name = Event.NAME_DONOR_URINE2
        urine_event_2.save()

        urine_event_2_sample_1 = UrineSample()
        urine_event_2_sample_1.person = person
        urine_event_2_sample_1.event = urine_event_2
        urine_event_2_sample_1.save()

    for kidney in (donor.left_kidney, donor.right_kidney):
        # ==========================================================  PERFUSATE 1
        perfusate_event_1 = Event()
        perfusate_event_1.type = Event.TYPE_PERFUSATE
        perfusate_event_1.name = Event.NAME_ORGAN_PERFUSATE1
        perfusate_event_1.save()

        perfusate_event_1_sample_1 = PerfusateSample()
        perfusate_event_1_sample_1.organ = kidney
        perfusate_event_1_sample_1.event = perfusate_event_1
        perfusate_event_1_sample_1.save()

        # ==========================================================  PERFUSATE 2
        perfusate_event_2 = Event()
        perfusate_event_2.type = Event.TYPE_PERFUSATE
        perfusate_event_2.name = Event.NAME_ORGAN_PERFUSATE2
        perfusate_event_2.save()

        perfusate_event_2_sample_1 = PerfusateSample()
        perfusate_event_2_sample_1.organ = kidney
        perfusate_event_2_sample_1.event = perfusate_event_2
        perfusate_event_2_sample_1.save()

        # ==========================================================  PERFUSATE 3
        perfusate_event_3 = Event()
        perfusate_event_3.type = Event.TYPE_PERFUSATE
        perfusate_event_3.name = Event.NAME_ORGAN_PERFUSATE3
        perfusate_event_3.save()

        perfusate_event_3_sample_1 = PerfusateSample()
        perfusate_event_3_sample_1.organ = kidney
        perfusate_event_3_sample_1.event = perfusate_event_3
        perfusate_event_3_sample_1.save()

        # ==========================================================  TISSUE 1
        tissue_event_1 = Event()
        tissue_event_1.type = Event.TYPE_TISSUE
        tissue_event_1.name = Event.NAME_ORGAN_TISSUE
        tissue_event_1.save()

        tissue_event_1_sample_1 = TissueSample()
        tissue_event_1_sample_1.organ = kidney
        tissue_event_1_sample_1.event = tissue_event_1
        tissue_event_1_sample_1.tissue_type = TissueSample.SAMPLE_R
        tissue_event_1_sample_1.save()

        tissue_event_1_sample_2 = TissueSample()
        tissue_event_1_sample_2.organ = kidney
        tissue_event_1_sample_2.event = tissue_event_1
        tissue_event_1_sample_2.tissue_type = TissueSample.SAMPLE_F
        tissue_event_1_sample_2.save()


@transaction.atomic
def create_recipient_samples(recipient):
    """
    For a given Recipient record, look to see if a corresponding Worksheet exists, and if not, create it
    and the related events and samples for the protocol
    :param recipient: Recipient instance
    :return: True, if created, False otherwise
    """
    person = recipient.person

    # ==========================================================  BLOOD 1
    blood_event_1 = Event()
    blood_event_1.type = Event.TYPE_BLOOD
    blood_event_1.name = Event.NAME_RECIPIENT_BLOOD1
    blood_event_1.save()

    blood_event_1_sample_1 = BloodSample()
    blood_event_1_sample_1.person = person
    blood_event_1_sample_1.event = blood_event_1
    blood_event_1_sample_1.blood_type = BloodSample.SAMPLE_SST
    blood_event_1_sample_1.save()

    blood_event_1_sample_2 = BloodSample()
    blood_event_1_sample_2.person = person
    blood_event_1_sample_2.event = blood_event_1
    blood_event_1_sample_2.blood_type = BloodSample.SAMPLE_EDSA
    blood_event_1_sample_2.save()

    # ==========================================================  BLOOD 2
    blood_event_2 = Event()
    blood_event_2.type = Event.TYPE_BLOOD
    blood_event_2.name = Event.NAME_RECIPIENT_BLOOD2
    blood_event_2.save()

    blood_event_2_sample_1 = BloodSample()
    blood_event_2_sample_1.person = person
    blood_event_2_sample_1.event = blood_event_2
    blood_event_2_sample_1.blood_type = BloodSample.SAMPLE_SST
    blood_event_2_sample_1.save()

    blood_event_2_sample_2 = BloodSample()
    blood_event_2_sample_2.person = person
    blood_event_2_sample_2.event = blood_event_2
    blood_event_2_sample_2.blood_type = BloodSample.SAMPLE_EDSA
    blood_event_2_sample_2.save()


def number_as_str(value):
    """
    Convert excel file values that appear as integers or floats, into string results that look like they do in the 
    application (i.e. 1234.0 in file -> 1234 as output)
    
    :param value:
    :return: string
    """
    output = ""
    try:
        output = int(value)
    except ValueError:
        try:
            output = int(float(value))
        except ValueError:  # "could not convert string to float"
            pass
    return str(output)


class WP7Workbook(Workbook):
    __workbook__ = None
    __current_row_id__ = 2  # First row is headers, so start with data
    __current_row_data__ = dict()  # Store the row in a dictionary based on the column titles
    __headers_by_id__ = dict()
    __headers_by_title__ = dict()

    def __init__(self, workbook=None):
        if workbook is not None:
            self.__workbook__ = workbook
            self.__load_header_dictionaries__()
        super(WP7Workbook, self).__init__()

    def __load_header_dictionaries__(self):
        # Find the headers in row 1 of this worksheet, and store in the first dictionary, and flip for the second
        for column_id in range(1, self.worksheet.max_column + 1):
            self.__headers_by_id__[column_id] = self.worksheet.cell(row=1, column=column_id).value
            self.__headers_by_title__[self.worksheet.cell(row=1, column=column_id).value] = column_id

    def load_xlsx(self, filename=None):
        if isinstance(filename, str) or isinstance(filename, IOBase):
            self.__workbook__ = load_workbook(filename)
            self.__load_header_dictionaries__()
            return True
        return False

    def load_row(self, row_id=0):
        """
        Load the specified row contents into a dictionary keyed on the column title. The title is
        reduced to lower case so that it matches model naming conventions and is consistent.
        :param row_id: int representing the worksheet row. Must be 2 or greater as row 1 is titles
        :return: dict() of row data if row id provided, otherwise, int of the current row id
        """
        if row_id > 1:
            self.__current_row_id__ = row_id
            for column_id in range(1, self.worksheet.max_column + 1):
                self.__current_row_data__[self.__headers_by_id__[column_id].lower()] = \
                    self.worksheet.cell(row=row_id, column=column_id).value
            return self.__current_row_data__
        return self.__current_row_id__

    def headers(self, key=None, value=None):
        if key is not None:
            return self.__headers_by_id__[key]
        if value is not None:
            return self.__headers_by_title__[value]
        return self.__headers_by_id__

    @property
    def workbook(self):
        return self.__workbook__

    @property
    def worksheet(self):
        return self.workbook.active


def get_sample_by_barcode(barcode):
    """
    Work through the 4 sample types to find one that has a barcode that matches
    :param barcode: 
    :return: 
    """
    if barcode is None or barcode == '':
        return None

    for result in BloodSample.objects.filter(barcode=barcode):
        # print("get_sample_by_barcode() match found with {0}".format(result))
        return result
    for result in UrineSample.objects.filter(barcode=barcode):
        # print("get_sample_by_barcode() match found with {0}".format(result))
        return result
    for result in TissueSample.objects.filter(barcode=barcode):
        # print("get_sample_by_barcode() match found with {0}".format(result))
        return result
    for result in PerfusateSample.objects.filter(barcode=barcode):
        # print("get_sample_by_barcode() match found with {0}".format(result))
        return result

    return None


def load_wp7_xlsx(workbook):
    """
    For a given workbook, create and update the WP7 samples record (used from either command line or UI)
    :param workbook: WP7Workbook object
    :return total_rows, created_count, linked_count: int, int, int - used as feedback to help track changes
    """
    from .forms import WP7RecordForm  # Done locally, else there's import recursion issues on loading
    from .models import WP7Record

    total_rows = workbook.worksheet.max_row
    created_count = 0
    linked_count = 0
    print("load_wp7_xlsx: Workbook is now loaded. {0} rows were found".format(total_rows))

    # As per Issue #296, wipe all existing records before importing again
    WP7Record.objects.all().delete()

    # Iterate through the data, creating or updating records in WP7Record
    for row_index in range(2, total_rows + 1):
        row_data = workbook.load_row(row_index)

        # print("DEBUG: load_wp7_xlsx: Row data for row {0} is loaded: {1}".format(row_index, row_data))

        data_form = WP7RecordForm(data={
            'barcode': row_data["scannedbarcode"],
            'box_number': row_data["boxnumber"],
            'position_in_box': row_data["positioninbox"]
        })
        if data_form.is_valid():
            if data_form.has_changed():
                record = data_form.save()
                created_count += 1
                record.content_object = get_sample_by_barcode(row_data["scannedbarcode"])
                if record.content_object is not None:
                    linked_count += 1
                record.save()
            # print("DEBUG: load_wp7_xlsx: record {0} has content_type {1} for matched_sample {2}".format(
            #     record, record.content_type, matched_sample))
        else:
            print("Form #{0} is INVALID".format(row_index))
            print(data_form.errors)

    return total_rows, created_count, linked_count
